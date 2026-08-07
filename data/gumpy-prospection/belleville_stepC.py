"""Step C : XLSX mairies rayon 50 km Belleville-sur-Saône."""
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

contacts = json.load(open("/opt/data/prospecting_mairies/belleville_contacts.json"))
df = pd.DataFrame(contacts).sort_values("distance_km")

# tranche de distance
def band(d):
    if d <= 15: return "0-15 km"
    if d <= 30: return "15-30 km"
    return "30-50 km"
df["tranche"] = df["distance_km"].apply(band)

try:
    events = json.load(open("/opt/data/prospecting_mairies/belleville_events.json"))
except FileNotFoundError:
    events = {}

wb = Workbook()
ws = wb.active
ws.title = "Mairies"
headers = ["Commune", "Département", "Population", "Distance (km)", "Tranche",
           "Email mairie", "Téléphone", "Site web", "Événements connus"]
ws.append(headers)
hdr_fill = PatternFill("solid", fgColor="1F4E78")
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = hdr_fill

for _, r in df.iterrows():
    ws.append([r["commune"], r["dept"], r["population"], r["distance_km"], r["tranche"],
               r["email"], r["tel"], r["site"], events.get(r["commune"], "")])

widths = [28, 12, 11, 13, 12, 32, 18, 34, 55]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# récap par tranche
ws2 = wb.create_sheet("Tranches")
ws2.append(["Tranche distance", "Nb communes", "Total population"])
for t, g in df.groupby("tranche", sort=False):
    ws2.append([t, len(g), g["population"].sum()])
for i, w in enumerate([18, 12, 18], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# typologie événements
ws3 = wb.create_sheet("Typologie événements")
ws3.append(["Période", "Événement type", "Note prospection"])
typ = [
    ("Déc", "Marché de Noël / Père Noël / illuminations", "Prospecter sept-oct"),
    ("Nov", "Fête du Beaujolais nouveau / fête des conscrits", "Prospecter sept-oct (spécifique Beaujolais)"),
    ("Oct-Nov", "Fête des aînés / repas des anciens", "Prospecter sept ; souvent CCAS"),
    ("Juin-Sept", "Fête du village / fête du pain / vide-greniers", "Prospecter mars-avril"),
    ("Sept", "Forum des associations", "Bonne porte d'entrée : présentiel"),
    ("Jan", "Cérémonie des vœux du maire", "Prospecter nov-déc"),
]
for row in typ:
    ws3.append(list(row))
for i, w in enumerate([12, 46, 40], start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

out = "/opt/data/prospecting_mairies/mairies_50km_belleville.xlsx"
wb.save(out)
print(f"Fichier : {out}")
print(f"{len(df)} communes | emails {df['email'].notna().sum()} | tél {df['tel'].notna().sum()}")
print("\nPar tranche :")
for t, g in df.groupby("tranche", sort=False):
    print(f"  {t}: {len(g)} communes")
