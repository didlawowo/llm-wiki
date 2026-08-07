"""Étape 3 : génération du fichier XLSX de mailing mairies 73+74."""
import json
import re
import unicodedata
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

contacts = json.load(open("/opt/data/prospecting_mairies/contacts.json"))

# --- fix Aix-les-Bains (email/tél officiels service-public) ---
for c in contacts:
    if c["commune"] == "Aix-les-Bains":
        c["email"] = "mairie@aixlesbains.fr"
        c["tel"] = "04 79 35 07 95"

# --- arrondissement depuis Wikipedia (segmentation secteur) ---
def norm(name: str) -> str:
    s = name.lower()
    s = s.replace("'", " ").replace("-", " ").replace("’", " ")
    s = re.sub(r"\(.*?\)", "", s)
    return re.sub(r"\s+", " ", s).strip()

arr = {}
for f in ["wiki_73.html", "wiki_74.html"]:
    for _, r in pd.read_html(f"/opt/data/prospecting_mairies/{f}", header=0)[1].iterrows():
        arr[norm(re.sub(r"\[.*?\]", "", str(r["Nom"])).strip())] = str(r["Arrondissement"])
for c in contacts:
    c["arrondissement"] = arr.get(norm(c["commune"]), "")

# --- secteur (arrondissement = zone de prospection) ---
for c in contacts:
    c["secteur"] = c["arrondissement"]

df = pd.DataFrame(contacts).sort_values(["secteur", "population"], ascending=[True, False])

# --- événements ---
try:
    events = json.load(open("/opt/data/prospecting_mairies/events.json"))
except FileNotFoundError:
    events = {}

# --- XLSX ---
wb = Workbook()
ws = wb.active
ws.title = "Mairies"

headers = ["Commune", "Département", "Population", "Secteur", "Arrondissement",
           "Email mairie", "Téléphone", "Site web", "Événements connus"]
ws.append(headers)
hdr_fill = PatternFill("solid", fgColor="1F4E78")
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = hdr_fill

for _, r in df.iterrows():
    ev = events.get(r["commune"], "")
    ws.append([r["commune"], f"73/74"[0:2], r["population"], r["secteur"],
               r["arrondissement"], r["email"], r["tel"], r["site"], ev])
    # corriger département proprement
    ws.cell(row=ws.max_row, column=2).value = r["dept"]

widths = [26, 12, 11, 14, 16, 32, 18, 34, 50]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# sheet secteurs récap
ws2 = wb.create_sheet("Secteurs")
ws2.append(["Secteur", "Nb communes", "Total population"])
for s, g in df.groupby("secteur"):
    ws2.append([s, len(g), g["population"].sum()])
for i, w in enumerate([18, 12, 18], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# sheet typologie événements (petites communes)
ws3 = wb.create_sheet("Typologie événements")
ws3.append(["Période", "Événement type", "Note prospection"])
typ = [
    ("Déc", "Marché de Noël / Père Noël / illuminations", "Prospecter sept-oct"),
    ("Oct-Nov", "Fête des aînés / repas des anciens", "Prospecter sept ; souvent organisée par le CCAS"),
    ("Juin-Sept", "Fête du village / fête du pain / vide-greniers", "Prospecter mars-avril"),
    ("Sept", "Forum des associations", "Bonne porte d'entrée : présentiel"),
    ("Jan", "Cérémonie des vœux du maire", "Prospecter nov-déc"),
    ("Toute l'année", "Fêtes de quartier, lotos, thés dansants", "Suivre l'agenda jds.fr de la commune"),
]
for row in typ:
    ws3.append(list(row))
for i, w in enumerate([12, 46, 40], start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

out = "/opt/data/prospecting_mairies/mairies_73_74_prospection.xlsx"
wb.save(out)
print(f"Fichier généré : {out}")
print(f"{len(df)} communes | emails {df['email'].notna().sum()} | tél {df['tel'].notna().sum()}")
print("\nSecteurs :")
for s, g in df.groupby("secteur"):
    print(f"  {s}: {len(g)} communes")
