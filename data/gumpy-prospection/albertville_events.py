"""XLSX : fêtes >7000 hab autour d'Albertville — organisateurs + contacts + sites de mise en valeur."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

rows = [
    # commune, événement, période, organisateur, email, téléphone, site organisateur, sites mise en valeur
    ("Albertville", "Festival International des Musiques Militaires (46e éd.)", "4-5 juillet",
     "Comité des Fêtes d'Albertville", "cdfalbertville@gmail.com", "06 23 78 04 80",
     "comitedesfetesalbertville.fr",
     "pays-albertville.com (agenda OT) · albertville.fr/evenements · jds.fr · savoie-news.fr"),
    ("Albertville", "Les Lumières de Noël : marché de Noël, Père Noël, concerts de l'Avent", "Décembre",
     "Ville d'Albertville (service animation)", "ville@albertville.fr", "04 79 10 43 00",
     "albertville.fr",
     "noel.org · jds.fr · eterritoire.fr · presse locale (Le Dauphiné, La Savoie)"),
    ("Bourg-Saint-Maurice", "Animations d'été : Transvanoise (moto), Fêtes de l'Edelweiss, des bergers, des myrtilles, de la terre", "Juin-septembre",
     "Office de Tourisme Haute Tarentaise / Les Arcs", "contact@lesarcs.com", "04 79 07 12 57",
     "hautetarentaise.fr · bourgsaintmaurice.fr/agenda",
     "planetekiosque.com · jds.fr · infolocale.fr · france-voyage.com"),
    ("Bourg-Saint-Maurice", "La Démontagnée (retour des troupeaux)", "Octobre",
     "Office de Tourisme Haute Tarentaise + mairie", "contact@lesarcs.com", "04 79 07 12 57",
     "hautetarentaise.fr",
     "planetekiosque.com · jds.fr · presse locale"),
    ("Bourg-Saint-Maurice", "Marché de Noël / animations hiver", "Décembre",
     "Mairie de Bourg-Saint-Maurice", "mairie@bourgsaintmaurice.fr", "04 79 07 23 33",
     "bourgsaintmaurice.fr",
     "noel.org · jds.fr"),
    ("Ugine", "Fête des Montagnes (58e éd. — défilé, terroir, associations)", "Septembre",
     "Ugine Animation (par délégation de la ville)", "ugine.animation@ugine.com", "06 29 02 34 29",
     "facebook.com/UgineAnimation",
     "fest.fr · planetekiosque.com · savoie-news.fr · La Savoie (Le Messager)"),
    ("Ugine", "Fête du Village + Fête de la Musique + foire locale", "Juin",
     "Ugine Animation + ville d'Ugine", "ugine.animation@ugine.com", "06 29 02 34 29",
     "ugine.com (site ville)",
     "fest.fr · jds.fr · planetekiosque.com"),
    ("Ugine", "Randonnée des Saveurs + animations estivales", "Juillet-août",
     "Office de Tourisme (programme à l'OT)", "ugine.animation@ugine.com", "04 79 37 33 00",
     "ugine.com · OT Arlysère",
     "france-voyage.com · jds.fr"),
]

wb = Workbook()
ws = wb.active
ws.title = "Fêtes & organisateurs"
headers = ["Commune", "Événement", "Période", "Organisateur", "Email", "Téléphone", "Site organisateur", "Sites de mise en valeur"]
ws.append(headers)
hdr = PatternFill("solid", fgColor="1F4E78")
for col in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = hdr
for r in rows:
    ws.append(list(r))
for i, w in enumerate([16, 55, 16, 38, 30, 16, 42, 70], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# sheet sites utiles
ws2 = wb.create_sheet("Sites utiles")
ws2.append(["Site", "Type", "Usage"])
sites = [
    ("jds.fr", "Agenda sorties par ville", "Rechercher les fêtes et les organisateurs"),
    ("planetekiosque.com", "Agenda animations/fêtes local", "Voir le détail des animations (dates, organisateur)"),
    ("infolocale.fr", "Agenda presse locale", "Annoncer / trouver les événements"),
    ("fest.fr", "Agenda des fêtes", "Fêtes de village, traditionnelles"),
    ("eterritoire.fr", "Agenda par commune", "Détail des manifestations locales"),
    ("france-voyage.com (événements)", "Agenda touristique", "Événements par commune"),
    ("noel.org", "Marchés de Noël", "Saison Noël"),
    ("pays-albertville.com", "Agenda OT Pays d'Albertville", "Tous les événements de la vallée"),
    ("hautetarentaise.fr", "OT Haute Tarentaise", "Animations Bourg-Saint-Maurice/Les Arcs"),
    ("albertville.fr / bourgsaintmaurice.fr / ugine.com", "Sites mairies (agenda)", "Agenda officiel des communes"),
    ("savoie-news.fr · La Savoie (Le Messager) · Le Dauphiné", "Presse locale", "Couverture des fêtes, photos presse"),
    ("Facebook (Comité des Fêtes, Ugine Animation)", "Réseaux sociaux", "Programmes, photos, actualités"),
]
for s in sites:
    ws2.append(list(s))
for i, w in enumerate([45, 30, 60], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

out = "/opt/data/prospecting_mairies/fetes_albertville_7000.xlsx"
wb.save(out)
print("Fichier :", out)
print(f"{len(rows)} événements, {len(sites)} sites utiles")
